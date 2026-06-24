"""Export (export-spec): inventory-kolommen, lege summary onder --no-llm, relations-graaf."""

import json

from openpyxl import load_workbook

from zeef.export import (
    INVENTORY_COLUMNS,
    write_criteria,
    write_inventory,
    write_manifest,
    write_relations,
)
from zeef.models import Criteria, Criterion, Document


def _doc(doc_id, **kw):
    d = Document(id=doc_id, source_path=f"/{doc_id}", doc_type=kw.get("doc_type", "email"))
    d.scores["final"] = kw.get("final", 0.5)
    d.decision = "selected"
    d.decision_reason = kw.get("reason", "top-n")
    d.rationale = kw.get("rationale", "")
    d.topic = kw.get("topic", "")
    d.subtopic = kw.get("subtopic", "")
    return d


def _cell(row, name):
    """Lees een cel op kolomnaam, niet -index (de category/doc_type-rebind wijzigt de volgorde)."""
    return row[INVENTORY_COLUMNS.index(name)]


def test_inventory_has_required_columns(tmp_path):
    docs = [_doc("a", final=0.91, reason="top-n=50", rationale="scoort hoog: publicatieclausule")]
    path = write_inventory(docs, tmp_path / "inventory.xlsx")
    ws = load_workbook(path).active
    assert tuple(c.value for c in ws[1]) == INVENTORY_COLUMNS
    row = [c.value for c in ws[2]]
    assert _cell(row, "id") == "a" and _cell(row, "score") == 0.91
    assert _cell(row, "reason") == "top-n=50"
    assert _cell(row, "motivatie") == "scoort hoog: publicatieclausule"


def test_category_is_topic_and_doc_type_kept_separate(tmp_path):
    docs = [_doc("a", doc_type="pdf_digital", topic="Subsidie cultuur", subtopic="Begroting 2026")]
    ws = load_workbook(write_inventory(docs, tmp_path / "inv.xlsx")).active
    row = [c.value for c in ws[2]]
    assert _cell(row, "category") == "Subsidie cultuur / Begroting 2026"
    assert _cell(row, "doc_type") == "pdf_digital"


def test_no_llm_leaves_motivatie_empty(tmp_path):
    ws = load_workbook(write_inventory([_doc("a")], tmp_path / "inv.xlsx")).active
    motivatie_idx = INVENTORY_COLUMNS.index("motivatie")
    assert [c.value for c in ws[2]][motivatie_idx] in ("", None)


def test_criteria_exported_as_json(tmp_path):
    crit = Criteria(query="begroting subsidie cultuur 2026",
                    items=[Criterion(label="onderwerp", description="gaat over de begroting")],
                    source="llm")
    path = write_criteria(crit, tmp_path / "criteria.json")
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["query"] == "begroting subsidie cultuur 2026"
    assert data["source"] == "llm"
    assert data["items"][0]["label"] == "onderwerp"


def test_no_llm_leaves_summary_empty(tmp_path):
    docs = [_doc("a")]  # geen metadata['summary'] gezet
    ws = load_workbook(write_inventory(docs, tmp_path / "inv.xlsx")).active
    summary_idx = INVENTORY_COLUMNS.index("summary")
    assert [c.value for c in ws[2]][summary_idx] in ("", None)


def test_manifest_exported_as_json(tmp_path):
    manifest = {
        "schema": "zeef-run-manifest/1",
        "query": "begroting subsidie cultuur 2026",
        "runtime_ms": {"total": 12.3, "stages": [{"stage": "ingest", "elapsed_ms": 4.2}]},
    }
    path = write_manifest(manifest, tmp_path / "run-manifest.json")
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["schema"] == "zeef-run-manifest/1"
    assert data["runtime_ms"]["total"] == 12.3
    assert data["runtime_ms"]["stages"][0]["stage"] == "ingest"


def test_relations_exported_as_graph(tmp_path):
    a = Document(id="a", source_path="/a", doc_type="email")
    b = Document(id="b", source_path="/b", doc_type="email")
    b.add_relation("duplicate-of", "a", evidence="sha256:deadbeef")
    import json

    path = write_relations([a, b], tmp_path / "relations.json")
    data = json.loads(path.read_text())
    assert data["edge_count"] == 1
    edge = data["edges"][0]
    assert edge == {"source": "b", "kind": "duplicate-of", "target": "a",
                    "evidence": "sha256:deadbeef"}
