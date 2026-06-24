"""End-to-end acceptatie (cli-spec): `sovereign --no-llm` op gemengde .eml/PDF, zónder netwerk.

Dekt de change #1-acceptatiecriteria: drie artefacten geproduceerd; een 5-mail thread valt
samen tot één eenheid; een exact duplicaat bezet één slot; elk uitgesloten document draagt een
reden in de audit-log. Een fixture blokkeert uitgaande sockets, zodat 'air-gapped' echt wordt
afgedwongen i.p.v. aangenomen.
"""

import json

import pytest

from zeef.audit import AuditLog
from zeef.config import CutoffMode, ProfileName, Settings
from zeef.pipeline.run import run_converge
from zeef.profiles import resolve_providers

QUERY = "begroting subsidie cultuur 2026"


@pytest.fixture
def no_network(monkeypatch):
    import socket

    def guard(*args, **kwargs):
        raise OSError("netwerk uitgeschakeld in air-gapped acceptatietest")

    monkeypatch.setattr(socket.socket, "connect", guard)
    monkeypatch.setattr(socket.socket, "connect_ex", guard)
    monkeypatch.setattr(socket, "create_connection", guard)


def _run(corpus, out_dir):
    settings = Settings(_env_file=None)
    providers = resolve_providers(ProfileName.sovereign, no_llm=True, settings=settings)
    audit = AuditLog(out_dir / "audit.jsonl")
    result = run_converge(corpus, QUERY, providers, CutoffMode.target, 100, out_dir, audit)
    return result, audit


def _by_name(result):
    return {d.source_path.split("/")[-1].split("#")[0]: d for d in result.documents}


def test_three_artifacts_present(corpus, tmp_path, no_network):
    result, _ = _run(corpus, tmp_path)
    for name in ("inventory.xlsx", "relations.json", "audit.jsonl"):
        assert (tmp_path / name).exists(), name
    assert result.counts()["selected"] > 0


def test_no_summary_column_in_inventory_under_no_llm(corpus, tmp_path, no_network):
    # Legt de bedrading vast: run.py geeft onder --no-llm `include_summary=False` door, dus de
    # geëxporteerde inventory heeft géén summary-kolom (geen lege kolom met header).
    from openpyxl import load_workbook

    _run(corpus, tmp_path)
    ws = load_workbook(tmp_path / "inventory.xlsx").active
    header = [c.value for c in ws[1]]
    assert "summary" not in header
    assert "category" in header and "motivatie" in header


def test_run_manifest_records_stage_runtimes(corpus, tmp_path, no_network):
    result, _ = _run(corpus, tmp_path)
    path = tmp_path / "run-manifest.json"
    assert path.exists()
    manifest = json.loads(path.read_text(encoding="utf-8"))
    assert manifest["schema"] == "zeef-run-manifest/1"
    assert manifest["query"] == QUERY
    assert isinstance(manifest["runtime_ms"]["total"], (int, float))
    stages = {s["stage"] for s in manifest["runtime_ms"]["stages"]}
    expected = {"criteria", "ingest", "relate", "scope-filter",
                "retrieve", "rerank", "score", "select", "export"}
    assert expected <= stages
    assert all(isinstance(s["elapsed_ms"], (int, float)) for s in manifest["runtime_ms"]["stages"])
    # Het manifest moet ook op het RunResult beschikbaar zijn voor de CLI-samenvatting.
    assert result.manifest is not None and result.manifest["runtime_ms"]["stages"]


def test_five_mail_thread_collapses_to_one_unit(corpus, tmp_path, no_network):
    result, _ = _run(corpus, tmp_path)
    by_name = _by_name(result)
    thread = [by_name[f"thread-0{i}.eml"] for i in range(1, 6)]
    selected_in_thread = [d for d in thread if d.decision == "selected"]
    assert len(selected_in_thread) == 1
    assert selected_in_thread[0] is by_name["thread-05.eml"]
    assert all(d.decision == "out_of_scope" for d in thread if d is not by_name["thread-05.eml"])


def test_exact_duplicate_occupies_one_slot(corpus, tmp_path, no_network):
    result, _ = _run(corpus, tmp_path)
    by_name = _by_name(result)
    assert by_name["dup-b.eml"].decision == "out_of_scope"
    assert by_name["dup-a.eml"].decision != "out_of_scope"


def test_every_excluded_doc_has_reason_in_audit(corpus, tmp_path, no_network):
    result, audit = _run(corpus, tmp_path)
    events = [json.loads(line) for line in audit.path.read_text().splitlines()]
    excluded_ids = {d.id for d in result.documents if d.decision == "out_of_scope"}
    reasoned = {e["document_ids"][0] for e in events
                if e["action"] == "excluded" and e["inputs"].get("reason")}
    assert excluded_ids and excluded_ids <= reasoned
    # En op het model zelf draagt elk uitgesloten document een niet-lege reden.
    assert all(d.decision_reason.strip() for d in result.documents
               if d.decision == "out_of_scope")


def test_audit_has_event_per_stage(corpus, tmp_path, no_network):
    _, audit = _run(corpus, tmp_path)
    stages = {json.loads(line)["stage"] for line in audit.path.read_text().splitlines()}
    assert {"ingest", "relate", "scope-filter", "retrieve", "rerank", "select", "export"} <= stages
